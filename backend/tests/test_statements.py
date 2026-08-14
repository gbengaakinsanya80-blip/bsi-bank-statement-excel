"""Financial statement tests: income statement, balance sheet, cash flow.

Phase 4 of the FinancePilot AI build: reports are derived from posted
journals via the trial balance so every number ties back to the books.
The cash flow uses the indirect method and must reconcile to the actual
movement in the cash accounts.
"""

from __future__ import annotations

import pytest

from app.accounting import posting, report_pdf, report_xlsx, statements
from app.accounting.coa import generate_default_coa
from app.export.sqlite_store import Store


@pytest.fixture()
def store() -> Store:
    return Store()


@pytest.fixture()
def company(store: Store) -> dict:
    company = store.create_company(user_id="u1", name="Ajao Traders", industry="retail")
    store.replace_chart_of_accounts(company["id"], generate_default_coa("retail"))
    return company


def _apply_txn(store: Store, company_id: str, statement_id: str = "s1", **fields) -> dict:
    row = {
        "row_index": 0,
        "tx_date": "2026-01-05",
        "description": "TEST TXN",
        "reference": "R1",
        "debit": None,
        "credit": None,
        "balance": 0,
        "category": "Other",
        "account_code": "1200",
        "transaction_type": "Debit",
        "confidence": 0.95,
        "rationale": "test",
        "status": "applied",
        "source": "categorizer",
    }
    row.update(fields)
    store.import_ledger_transactions(
        company_id=company_id, statement_id=statement_id, job_id="j1", rows=[row]
    )
    return store.list_ledger_transactions(company_id, status="applied")[-1]


def _post_base_book(store: Store, company_id: str) -> None:
    """A profitable month: sales, rent, salary, a loan and equipment."""
    _apply_txn(
        store, company_id, description="CUSTOMER PAYMENT",
        credit=300000, account_code="4010", transaction_type="Credit",
    )
    _apply_txn(
        store, company_id, description="RENT", debit=100000, account_code="6020"
    )
    _apply_txn(
        store, company_id, description="SALARY", debit=50000, account_code="6010"
    )
    _apply_txn(
        store, company_id, description="LOAN RECEIVED",
        credit=200000, account_code="2060", transaction_type="Credit",
    )
    _apply_txn(
        store, company_id, description="EQUIPMENT", debit=80000, account_code="1070"
    )
    summary = posting.post_applied_transactions(
        store, company_id=company_id, user_id="u1"
    )
    assert summary["posted"] == 5


# ---------------------------------------------------------------------- #
# Income statement
# ---------------------------------------------------------------------- #
def test_income_statement_summarises_revenue_expenses_and_profit(
    store: Store, company: dict
) -> None:
    _post_base_book(store, company["id"])

    report = statements.income_statement(store, company_id=company["id"])

    assert report["total_revenue"] == 300000
    assert report["total_expenses"] == 150000
    assert report["net_profit"] == 150000
    assert {l["code"]: l["balance"] for l in report["revenue"]} == {"4010": 300000}
    assert {l["code"]: l["balance"] for l in report["expenses"]} == {
        "6010": 50000, "6020": 100000,
    }


# ---------------------------------------------------------------------- #
# Balance sheet
# ---------------------------------------------------------------------- #
def test_balance_sheet_balances_assets_liabilities_and_equity(
    store: Store, company: dict
) -> None:
    _post_base_book(store, company["id"])

    report = statements.balance_sheet(store, company_id=company["id"])

    assert report["total_assets"] == 350000  # Bank 270,000 + P&E 80,000
    assert report["total_liabilities"] == 200000  # Loans
    assert report["current_year_profit"] == 150000
    assert report["balancing_figure"] == 0.0
    assert report["total_equity"] == 150000
    assert report["balanced"] is True
    assert {l["code"]: l["balance"] for l in report["assets"]} == {
        "1010": 270000, "1070": 80000,
    }


def test_balance_sheet_keeps_the_report_balanced_even_with_unbalanced_books(
    store: Store, company: dict
) -> None:
    # A lone credit posted straight to the ledger (bypassing the engine) makes
    # the trial balance unbalanced, but the report still reconciles.
    store.create_journal_entry(
        company["id"],
        journal_no="JRNL-202601-9000",
        tx_date="2026-01-31",
        reference="TEST",
        description="Unbalanced",
        status="posted",
        source_type="manual",
        created_by="u1",
        lines=[{"account_code": "4010", "debit": 0, "credit": 1000}],
    )
    tb = posting.trial_balance(store, company_id=company["id"])
    assert tb["balanced"] is False

    report = statements.balance_sheet(store, company_id=company["id"])

    assert report["total_assets"] == 0
    assert report["current_year_profit"] == 1000
    assert report["balancing_figure"] == -1000
    assert report["total_equity"] == 0
    assert report["balanced"] is True


# ---------------------------------------------------------------------- #
# Cash flow
# ---------------------------------------------------------------------- #
def test_cash_flow_indirect_method_reconciles_to_bank_movement(
    store: Store, company: dict
) -> None:
    _post_base_book(store, company["id"])

    report = statements.cash_flow_statement(store, company_id=company["id"])

    assert report["operating"]["net_profit"] == 150000
    assert report["operating"]["net_cash"] == 150000
    assert report["investing"]["net_cash"] == -80000
    assert report["financing"]["net_cash"] == 200000
    assert report["net_increase_in_cash"] == 270000
    assert report["closing_cash"] == 270000
    assert report["ties_to_cash"] is True

    # P&E sits in investing, the loan in financing.
    investing_codes = {i["code"] for i in report["investing"]["items"]}
    financing_codes = {i["code"] for i in report["financing"]["items"]}
    assert investing_codes == {"1070"}
    assert financing_codes == {"2060"}


def test_accrual_adjustment_is_an_operating_non_cash_item(
    store: Store, company: dict
) -> None:
    _post_base_book(store, company["id"])
    adj = posting.create_adjustment(
        store, company_id=company["id"], user_id="u1",
        description="Accrue rent", amount=25000,
    )
    posting.approve_adjustment(
        store, company_id=company["id"], user_id="u1", adj_id=adj["id"],
        debit_code="6020", credit_code="2020",
    )

    income = statements.income_statement(store, company_id=company["id"])
    balance = statements.balance_sheet(store, company_id=company["id"])
    cash = statements.cash_flow_statement(store, company_id=company["id"])

    # The accrual reduces profit but adds the same amount back as a
    # working-capital adjustment: operating cash is unchanged.
    assert income["net_profit"] == 125000
    assert balance["total_liabilities"] == 225000
    assert balance["balanced"] is True
    assert cash["operating"]["net_profit"] == 125000
    adjustments = {a["code"]: a["change"] for a in cash["operating"]["adjustments"]}
    assert adjustments["2020"] == 25000
    assert cash["operating"]["net_cash"] == 150000
    assert cash["net_increase_in_cash"] == 270000
    assert cash["ties_to_cash"] is True


def test_suspense_credit_shows_as_operating_inflow(
    store: Store, company: dict
) -> None:
    # An uncleared transfer credited to Suspense (1200) is real cash in.
    _apply_txn(
        store, company["id"], description="TRANSFER IN",
        credit=100000, account_code="1200", transaction_type="Credit",
    )
    posting.post_applied_transactions(store, company_id=company["id"], user_id="u1")

    cash = statements.cash_flow_statement(store, company_id=company["id"])

    adjustments = {a["code"]: a["change"] for a in cash["operating"]["adjustments"]}
    assert adjustments["1200"] == 100000
    assert cash["operating"]["net_cash"] == 100000
    assert cash["net_increase_in_cash"] == 100000
    assert cash["ties_to_cash"] is True


# ---------------------------------------------------------------------- #
# Period scoping
# ---------------------------------------------------------------------- #
def test_reports_can_be_scoped_to_a_single_period(
    store: Store, company: dict
) -> None:
    jan = store.create_period(
        company["id"], name="Jan 2026",
        start_date="2026-01-01", end_date="2026-01-31",
    )
    feb = store.create_period(
        company["id"], name="Feb 2026",
        start_date="2026-02-01", end_date="2026-02-28",
    )
    jan_stmt = store.link_statement(
        company_id=company["id"], user_id="u1", job_id="j1", period_id=jan["id"]
    )
    feb_stmt = store.link_statement(
        company_id=company["id"], user_id="u1", job_id="j1", period_id=feb["id"]
    )
    _apply_txn(
        store, company["id"], statement_id=jan_stmt["id"], description="CUSTOMER PAYMENT",
        credit=300000, account_code="4010", transaction_type="Credit",
    )
    _apply_txn(
        store, company["id"], statement_id=feb_stmt["id"], description="SALARY",
        debit=100000, account_code="6010",
    )
    summary = posting.post_applied_transactions(
        store, company_id=company["id"], user_id="u1"
    )
    assert summary["posted"] == 2

    jan_report = statements.income_statement(
        store, company_id=company["id"], period_id=jan["id"]
    )
    feb_report = statements.income_statement(
        store, company_id=company["id"], period_id=feb["id"]
    )
    full = statements.income_statement(store, company_id=company["id"])

    assert jan_report["total_revenue"] == 300000
    assert jan_report["net_profit"] == 300000
    assert feb_report["total_expenses"] == 100000
    assert feb_report["net_profit"] == -100000
    assert full["net_profit"] == 200000

    jan_balance = statements.balance_sheet(
        store, company_id=company["id"], period_id=jan["id"]
    )
    assert jan_balance["total_assets"] == 300000
    assert jan_balance["balanced"] is True


# ---------------------------------------------------------------------- #
# PDF report rendering
# ---------------------------------------------------------------------- #
def test_report_pdf_renders_each_statement(store: Store, company: dict) -> None:
    _post_base_book(store, company["id"])
    builders = {
        "income-statement": statements.income_statement,
        "balance-sheet": statements.balance_sheet,
        "cash-flow": statements.cash_flow_statement,
    }
    for kind, builder in builders.items():
        data = builder(store, company_id=company["id"])
        pdf = report_pdf.build_report_pdf(
            company=company, report_kind=kind, data=data
        )
        assert pdf.startswith(b"%PDF")
        text = _pdf_text(pdf)
        assert report_pdf.REPORT_TITLES[kind] in text
        assert "Ajao Traders" in text
        assert "150,000.00" in text


def _pdf_text(pdf: bytes) -> str:
    import io

    import fitz

    with fitz.open(stream=pdf, filetype="pdf") as doc:
        return "\n".join(page.get_text() for page in doc)


def test_report_xlsx_renders_each_statement(store: Store, company: dict) -> None:
    import io

    from openpyxl import load_workbook

    _post_base_book(store, company["id"])
    builders = {
        "income-statement": statements.income_statement,
        "balance-sheet": statements.balance_sheet,
        "cash-flow": statements.cash_flow_statement,
    }
    for kind, builder in builders.items():
        data = builder(store, company_id=company["id"])
        raw = report_xlsx.build_report_xlsx(company=company, report_kind=kind, data=data)
        wb = load_workbook(io.BytesIO(raw))
        ws = wb.active
        assert ws.title == report_pdf.REPORT_TITLES[kind][:31]
        # Company name + title + period label precede the column header.
        assert ws.cell(row=1, column=1).value == "Ajao Traders"
        assert ws.cell(row=2, column=1).value == report_pdf.REPORT_TITLES[kind]
        values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        expected = {
            "income-statement": ["Net Profit", "Total Revenue", "Total Expenses"],
            "balance-sheet": ["Total Equity", "Total Assets", "Current Year Profit"],
            "cash-flow": ["Closing Cash", "Net Cash from Operating Activities"],
        }[kind]
        for label in expected:
            assert label in values


# ---------------------------------------------------------------------- #
# API layer
# ---------------------------------------------------------------------- #
@pytest.fixture(scope="module")
def client():
    from fastapi.testclient import TestClient

    from app.main import app

    return TestClient(app)


def _auth_headers(client, email: str) -> dict:
    r = client.post("/api/auth/register", json={"email": email, "password": "correct-horse-battery"})
    if r.status_code == 409:
        r = client.post("/api/auth/login", json={"email": email, "password": "correct-horse-battery"})
    assert r.status_code == 200, r.text
    return {"Authorization": f"Bearer {r.json()['token']}"}


@pytest.fixture(scope="module")
def post_headers(client) -> dict:
    return _auth_headers(client, "reports@bsi.local")


def test_api_report_endpoints(client, post_headers) -> None:
    import time
    import pathlib
    import tempfile

    from tests.generators.make_statements import build_statement_pdf

    r = client.post(
        "/api/companies", json={"name": "Report Co", "industry": "retail"}, headers=post_headers
    )
    assert r.status_code == 200, r.text
    company_id = r.json()["id"]

    period = client.post(
        f"/api/companies/{company_id}/periods",
        json={"name": "Jan 2026", "start_date": "2026-01-01", "end_date": "2026-01-31"},
        headers=post_headers,
    ).json()

    tmp = pathlib.Path(tempfile.mkdtemp())
    pdf = tmp / "stmt.pdf"
    build_statement_pdf(pdf, bank="Zenith Bank", n_transactions=8, seed=17)
    with pdf.open("rb") as fh:
        r = client.post(
            "/api/process", files={"upload": ("stmt.pdf", fh, "application/pdf")},
            headers=post_headers,
        )
    assert r.status_code == 200
    job_id = r.json()["job_id"]
    for _ in range(100):
        status = client.get(f"/api/jobs/{job_id}", headers=post_headers).json()["status"]
        if status in ("completed", "failed"):
            break
        time.sleep(0.2)
    assert status == "completed"

    link = client.post(
        f"/api/companies/{company_id}/statements",
        json={"job_id": job_id, "period_id": period["id"]},
        headers=post_headers,
    )
    assert link.status_code == 200, link.text
    statement_id = link.json()["id"]
    classify = client.post(
        f"/api/companies/{company_id}/statements/{statement_id}/classify", headers=post_headers
    )
    assert classify.status_code == 200, classify.text

    queue = client.get(f"/api/companies/{company_id}/classifications", headers=post_headers).json()
    for txn in queue["transactions"]:
        client.post(
            f"/api/companies/{company_id}/classifications/{txn['id']}/approve",
            headers=post_headers,
        )
    run = client.post(
        f"/api/companies/{company_id}/posting/run",
        json={"period_id": period["id"]},
        headers=post_headers,
    )
    assert run.status_code == 200, run.text
    assert run.json()["posted"] == queue["total"]

    base = f"/api/companies/{company_id}/reports"
    params = {"period_id": period["id"]}

    income = client.get(f"{base}/income-statement", params=params, headers=post_headers)
    assert income.status_code == 200, income.text
    body = income.json()
    assert "net_profit" in body
    assert body["total_revenue"] >= 0
    assert body["net_profit"] == round(body["total_revenue"] - body["total_expenses"], 2)

    balance = client.get(f"{base}/balance-sheet", params=params, headers=post_headers)
    assert balance.status_code == 200, balance.text
    sheet = balance.json()
    assert sheet["balanced"] is True
    assert sheet["total_assets"] == sheet["total_liabilities"] + sheet["total_equity"]

    cash = client.get(f"{base}/cash-flow", params=params, headers=post_headers)
    assert cash.status_code == 200, cash.text
    flow = cash.json()
    assert flow["ties_to_cash"] is True
    assert (
        flow["operating"]["net_cash"] + flow["investing"]["net_cash"]
        + flow["financing"]["net_cash"]
    ) == flow["net_increase_in_cash"]

    # Security: an unauthenticated request is refused.
    assert client.get(f"{base}/income-statement", params=params).status_code == 401


def test_api_report_pdf_download(client, post_headers) -> None:
    r = client.get(
        f"/api/companies/does-not-exist/reports/income-statement/pdf",
        headers=post_headers,
    )
    assert r.status_code == 404

    companies = client.get("/api/companies", headers=post_headers).json()["companies"]
    assert companies
    company_id = companies[0]["id"]

    for kind in ("income-statement", "balance-sheet", "cash-flow"):
        r = client.get(
            f"/api/companies/{company_id}/reports/{kind}/pdf", headers=post_headers
        )
        assert r.status_code == 200, r.text
        assert r.headers["content-type"] == "application/pdf"
        assert r.content.startswith(b"%PDF")
        assert "attachment" in r.headers["content-disposition"]

        x = client.get(
            f"/api/companies/{company_id}/reports/{kind}/xlsx", headers=post_headers
        )
        assert x.status_code == 200, x.text
        assert "spreadsheetml.sheet" in x.headers["content-type"]
        assert x.content.startswith(b"PK")
        assert "attachment" in x.headers["content-disposition"]

    bad = client.get(
        f"/api/companies/{company_id}/reports/nonsense/pdf", headers=post_headers
    )
    assert bad.status_code == 400
    bad_x = client.get(
        f"/api/companies/{company_id}/reports/nonsense/xlsx", headers=post_headers
    )
    assert bad_x.status_code == 400
