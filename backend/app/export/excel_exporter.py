"""Excel exporter.

Generates a professional workbook with four worksheets:

  1. Transactions  — full extracted ledger
  2. Summary       — opening/closing, totals, largest & average amounts
  3. Validation    — missing rows, balance errors, duplicates, unreadable rows
  4. Charts        — monthly cash flow, daily cash flow, income vs expense,
                     running balance trend
"""

from __future__ import annotations

import io
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..core.models import ParsedStatement, SummaryStats, Transaction, ValidationReport

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SUB_FONT = Font(bold=True, size=11, color="404040")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALT_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
DEBIT_FILL = PatternFill(start_color="FDECEA", end_color="FDECEA", fill_type="solid")
CREDIT_FILL = PatternFill(start_color="EAF6EC", end_color="EAF6EC", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFF7E6", end_color="FFF7E6", fill_type="solid")

TRANSACTION_COLUMNS = [
    ("Date", 12),
    ("Value Date", 12),
    ("Description", 48),
    ("Reference", 22),
    ("Debit", 14),
    ("Credit", 14),
    ("Balance", 16),
    ("Currency", 9),
    ("Branch", 14),
    ("Channel", 12),
    ("Instrument No", 14),
    ("Transaction Type", 15),
    ("Page", 7),
    ("Line", 7),
]


def _style_header(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _money(cell) -> None:
    cell.number_format = "#,##0.00"


def export_excel(parsed: ParsedStatement) -> bytes:
    wb = Workbook()

    _write_transactions(wb, parsed)
    _write_summary(wb, parsed.summary, parsed.meta.bank_name)
    _write_validation(wb, parsed.validation)
    _write_charts(wb, parsed)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# --------------------------------------------------------------------------- #
def _write_transactions(wb: Workbook, parsed: ParsedStatement) -> None:
    ws = wb.active
    ws.title = "Transactions"

    ws["A1"] = f"{parsed.meta.bank_name or 'Bank'} Account Statement — {parsed.meta.account_number or ''}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Period: {parsed.meta.period_start or ''} to {parsed.meta.period_end or ''}  |  Extraction: {parsed.meta.extraction_method}"
    ws["A2"].font = SUB_FONT

    header_row = 4
    for col_idx, (name, width) in enumerate(TRANSACTION_COLUMNS, start=1):
        ws.cell(row=header_row, column=col_idx, value=name)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    _style_header(ws, header_row, len(TRANSACTION_COLUMNS))

    def row_values(tx: Transaction) -> list[Any]:
        return [
            tx.tx_date.strftime("%d/%m/%Y") if tx.tx_date else "",
            tx.value_date.strftime("%d/%m/%Y") if tx.value_date else "",
            tx.description,
            tx.reference,
            tx.debit,
            tx.credit,
            tx.balance,
            tx.currency,
            tx.branch,
            tx.channel,
            tx.instrument_number,
            tx.tx_type,
            tx.page_number,
            tx.line_number,
        ]

    r = header_row + 1
    for tx in parsed.transactions:
        vals = row_values(tx)
        for col_idx, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col_idx, value=v)
            cell.border = BORDER
            if col_idx in (5, 6, 7) and isinstance(v, (int, float)):
                _money(cell)
        # Highlight balance records
        if tx.is_beginning_balance or tx.is_ending_balance:
            for col_idx in range(1, len(TRANSACTION_COLUMNS) + 1):
                ws.cell(row=r, column=col_idx).fill = ALT_FILL
                ws.cell(row=r, column=col_idx).font = Font(bold=True)
        elif tx.is_estimated:
            for col_idx in range(1, len(TRANSACTION_COLUMNS) + 1):
                ws.cell(row=r, column=col_idx).fill = WARN_FILL
        if tx.debit is not None:
            ws.cell(row=r, column=5).fill = DEBIT_FILL
        if tx.credit is not None:
            ws.cell(row=r, column=6).fill = CREDIT_FILL
        r += 1

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{header_row}:N{r - 1}"


# --------------------------------------------------------------------------- #
def _write_summary(wb: Workbook, s: SummaryStats, bank_name: str) -> None:
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Statement Summary"
    ws["A1"].font = TITLE_FONT

    rows: list[tuple[str, Any]] = [
        ("Opening Balance", s.opening_balance),
        ("Closing Balance", s.closing_balance),
        ("Total Credits (Pay In)", s.total_credits),
        ("Total Debits (Pay Out)", s.total_debits),
        ("Number of Transactions", s.number_of_transactions),
        ("Largest Debit", s.largest_debit),
        ("Largest Credit", s.largest_credit),
        ("Average Debit", s.average_debit),
        ("Average Credit", s.average_credit),
        ("Credit Transaction Count", s.total_credit_count),
        ("Debit Transaction Count", s.total_debit_count),
        ("Currency", s.currency),
    ]

    r = 3
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=r, column=2, value=value)
        if isinstance(value, (int, float)):
            _money(cell)
        r += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    # Monthly cash flow table
    start = r + 2
    ws.cell(row=start, column=1, value="Monthly Cash Flow").font = SUB_FONT
    r = start + 1
    for col_idx, name in enumerate(["Month", "Credits", "Debits", "Net"], start=1):
        ws.cell(row=r, column=col_idx, value=name)
    _style_header(ws, r, 4)
    r += 1
    for m in s.monthly_cash_flow:
        ws.cell(row=r, column=1, value=m["month"])
        ws.cell(row=r, column=2, value=m["credits"])
        ws.cell(row=r, column=3, value=m["debits"])
        ws.cell(row=r, column=4, value=m["net"])
        for c in range(2, 5):
            _money(ws.cell(row=r, column=c))
        r += 1


# --------------------------------------------------------------------------- #
def _write_validation(wb: Workbook, v: ValidationReport) -> None:
    ws = wb.create_sheet("Validation")
    ws["A1"] = "Extraction Validation Report"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        f"Balance reconciled: {'YES' if v.balance_reconciled else 'NO'}   "
        f"OCR confidence: {f'{v.ocr_confidence:.1%}' if v.ocr_confidence is not None else 'n/a'}"
    )
    ws["A2"].font = SUB_FONT

    sections = [
        ("Missing Rows", v.missing_rows),
        ("Balance Errors", v.balance_errors),
        ("Duplicate Entries", v.duplicate_entries),
        ("Unreadable Transactions", v.unreadable_transactions),
        ("Other Issues", v.other_issues),
    ]

    r = 4
    for title, issues in sections:
        ws.cell(row=r, column=1, value=f"{title} ({len(issues)})").font = Font(bold=True, size=12, color="1F4E79")
        r += 1
        for col_idx, name in enumerate(["Severity", "Page", "Line", "Message", "Expected", "Actual", "Suggested Fix"], start=1):
            ws.cell(row=r, column=col_idx, value=name)
        _style_header(ws, r, 7)
        r += 1
        if not issues:
            ws.cell(row=r, column=1, value="None").font = Font(italic=True, color="808080")
            r += 2
            continue
        for issue in issues:
            ws.cell(row=r, column=1, value=issue.severity)
            ws.cell(row=r, column=2, value=issue.page_number)
            ws.cell(row=r, column=3, value=issue.line_number)
            ws.cell(row=r, column=4, value=issue.message)
            ws.cell(row=r, column=5, value=issue.expected)
            ws.cell(row=r, column=6, value=issue.actual)
            ws.cell(row=r, column=7, value=issue.suggested_fix)
            if issue.severity == "error":
                for c in range(1, 8):
                    ws.cell(row=r, column=c).fill = WARN_FILL
            r += 1
        r += 1

    for col, width in [(1, 12), (2, 8), (3, 8), (4, 70), (5, 14), (6, 14), (7, 60)]:
        ws.column_dimensions[get_column_letter(col)].width = width


# --------------------------------------------------------------------------- #
def _write_charts(wb: Workbook, parsed: ParsedStatement) -> None:
    ws = wb.create_sheet("Charts")
    data = parsed.summary

    def write_table(sheet, start_row: int, header: list[str], rows: list[dict[str, Any]], keys: list[str]) -> int:
        for col_idx, name in enumerate(header, start=1):
            sheet.cell(row=start_row, column=col_idx, value=name)
        _style_header(sheet, start_row, len(header))
        r = start_row + 1
        for row in rows:
            for col_idx, key in enumerate(keys, start=1):
                sheet.cell(row=r, column=col_idx, value=row.get(key))
            r += 1
        return r - 1

    # Monthly cash flow chart
    r1 = write_table(ws, 1, ["Month", "Credits", "Debits"], data.monthly_cash_flow, ["month", "credits", "debits"])
    if r1 > 2:
        bar = BarChart()
        bar.type = "col"
        bar.title = "Monthly Cash Flow (Credits vs Debits)"
        bar.style = 10
        cats = Reference(ws, min_col=1, min_row=2, max_row=r1)
        c1 = Reference(ws, min_col=2, min_row=1, max_row=r1)
        c2 = Reference(ws, min_col=3, min_row=1, max_row=r1)
        bar.add_data(c1, titles_from_data=True)
        bar.add_data(c2, titles_from_data=True)
        bar.set_categories(cats)
        bar.height = 9
        bar.width = 20
        ws.add_chart(bar, "E2")

    # Daily cash flow
    top = 3 + max(len(data.monthly_cash_flow), 2) + 2
    r2 = write_table(ws, top, ["Day", "Credits", "Debits"], data.daily_cash_flow[:60], ["day", "credits", "debits"])
    if r2 > top + 1:
        line = LineChart()
        line.title = "Daily Cash Flow (Last 60 Days)"
        line.style = 12
        cats = Reference(ws, min_col=1, min_row=top + 1, max_row=r2)
        d1 = Reference(ws, min_col=2, min_row=top, max_row=r2)
        d2 = Reference(ws, min_col=3, min_row=top, max_row=r2)
        line.add_data(d1, titles_from_data=True)
        line.add_data(d2, titles_from_data=True)
        line.set_categories(cats)
        line.height = 9
        line.width = 20
        ws.add_chart(line, f"E{top}")

    # Income vs Expense
    top2 = top + max(len(data.daily_cash_flow[:60]), 2) + 2
    ws.cell(row=top2, column=1, value="Income vs Expense").font = SUB_FONT
    ws.cell(row=top2 + 1, column=1, value="Income").font = Font(bold=True)
    ws.cell(row=top2 + 1, column=2, value=data.total_credits)
    _money(ws.cell(row=top2 + 1, column=2))
    ws.cell(row=top2 + 2, column=1, value="Expense").font = Font(bold=True)
    ws.cell(row=top2 + 2, column=2, value=data.total_debits)
    _money(ws.cell(row=top2 + 2, column=2))
    pie = BarChart()
    pie.type = "bar"
    pie.title = "Income vs Expense"
    pie.add_data(Reference(ws, min_col=2, min_row=top2 + 1, max_row=top2 + 2), titles_from_data=False)
    pie.set_categories(Reference(ws, min_col=1, min_row=top2 + 1, max_row=top2 + 2))
    pie.height = 8
    pie.width = 14
    ws.add_chart(pie, f"E{top2}")
