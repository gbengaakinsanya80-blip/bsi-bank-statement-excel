"""Excel workbook export: Transactions, Summary, Validation, Insights and
Charts sheets with native Excel charts."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.models import InsightsReport, ParsedStatement, Transaction

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
ALT_FILL = PatternFill("solid", fgColor="EAF1F8")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = "#,##0.00"
MONEY_RED = "FF0000"

TRANSACTION_HEADERS = [
    "Beginning Balance", "Date", "Value Date", "Description", "Reference",
    "Debit", "Credit", "Balance", "Currency", "Branch", "Channel",
    "Instrument Number", "Transaction Type", "Category", "Page", "Line",
]


def build_excel(parsed: ParsedStatement) -> BytesIO:
    wb = Workbook()

    _write_transactions(wb.active, parsed)
    _write_summary(wb.create_sheet("Summary"), parsed)
    _write_account_heads(wb.create_sheet("Account Heads"), parsed)
    _write_validation(wb.create_sheet("Validation"), parsed)
    _write_insights(wb.create_sheet("Insights"), parsed)
    _write_charts(wb.create_sheet("Charts"), parsed)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def _style_header(ws, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        cell.border = BORDER


def _write_transactions(ws, parsed: ParsedStatement) -> None:
    ws.title = "Transactions"
    headers = TRANSACTION_HEADERS
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for i, t in enumerate(parsed.transactions, start=2):
        ws.append(_transaction_row(t))
        for col in range(1, len(headers) + 1):
            ws.cell(row=i, column=col).border = BORDER
            if i % 2 == 0:
                ws.cell(row=i, column=col).fill = ALT_FILL
        for col in (1, 6, 7, 8):
            ws.cell(row=i, column=col).number_format = MONEY
        if t.debit is not None:
            ws.cell(row=i, column=6).font = Font(color=MONEY_RED)

    _autosize(ws, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(parsed.transactions) + 1, 2)}"


def _transaction_row(t: Transaction) -> list:
    return [
        t.balance if t.is_beginning_balance else None,
        t.tx_date.strftime("%d/%m/%Y") if t.tx_date else None,
        t.value_date.strftime("%d/%m/%Y") if t.value_date else None,
        t.description,
        t.reference,
        t.debit,
        t.credit,
        t.balance,
        t.currency,
        t.branch,
        t.channel,
        t.instrument_number,
        t.tx_type,
        t.category,
        t.page_number,
        t.line_number,
    ]


def _write_summary(ws, parsed: ParsedStatement) -> None:
    s = parsed.summary
    meta = parsed.meta
    rows = [
        ("Bank", meta.bank_name),
        ("Account Name", meta.account_name),
        ("Account Number", meta.account_number),
        ("Period", f"{meta.period_start} to {meta.period_end}"),
        ("Currency", s.currency),
        ("", ""),
        ("Opening Balance", s.opening_balance),
        ("Closing Balance", s.closing_balance),
        ("Total Credits", s.total_credits),
        ("Total Debits", s.total_debits),
        ("Number of Transactions", s.number_of_transactions),
        ("Largest Debit", s.largest_debit),
        ("Largest Credit", s.largest_credit),
        ("Average Debit", s.average_debit),
        ("Average Credit", s.average_credit),
        ("Credit Count", s.total_credit_count),
        ("Debit Count", s.total_debit_count),
        ("Balance Reconciled", parsed.validation.balance_reconciled),
    ]
    for i, (label, value) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)
        if isinstance(value, float):
            ws.cell(row=i, column=2).number_format = MONEY

    # Monthly cash flow table (also feeds the Charts sheet references).
    start_row = len(rows) + 3
    ws.cell(row=start_row, column=1, value="Monthly Cash Flow").font = Font(bold=True, size=12)
    ws.cell(row=start_row + 1, column=1, value="Month")
    ws.cell(row=start_row + 1, column=2, value="Credits")
    ws.cell(row=start_row + 1, column=3, value="Debits")
    ws.cell(row=start_row + 1, column=4, value="Net")
    _style_header(ws, start_row + 1, 4)
    for i, m in enumerate(s.monthly_cash_flow, start=start_row + 2):
        ws.cell(row=i, column=1, value=m["month"])
        ws.cell(row=i, column=2, value=m["credits"]).number_format = MONEY
        ws.cell(row=i, column=3, value=m["debits"]).number_format = MONEY
        ws.cell(row=i, column=4, value=m["net"]).number_format = MONEY
    _autosize(ws, ["Statistic", "Value"])


def _write_account_heads(ws, parsed: ParsedStatement) -> None:
    headers = ["Account Head", "Transactions", "Debits", "Credits", "Net", "Share %"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    heads = parsed.account_heads
    total_turnover = sum(h.debit_total + h.credit_total for h in heads) or 0.0
    total_tx = sum(h.transaction_count for h in heads)
    for i, head in enumerate(heads, start=2):
        share = (100.0 * (head.debit_total + head.credit_total) / total_turnover) if total_turnover else 0.0
        ws.append([
            head.name,
            head.transaction_count,
            head.debit_total,
            head.credit_total,
            head.net,
            round(share, 2),
        ])
        for col in range(1, len(headers) + 1):
            ws.cell(row=i, column=col).border = BORDER
            if i % 2 == 0:
                ws.cell(row=i, column=col).fill = ALT_FILL
        for col in (3, 4, 5):
            ws.cell(row=i, column=col).number_format = MONEY
        ws.cell(row=i, column=6).number_format = "0.00"

    if heads:
        total_row = len(heads) + 2
        ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
        ws.cell(row=total_row, column=2, value=total_tx).font = Font(bold=True)
        ws.cell(row=total_row, column=3, value=sum(h.debit_total for h in heads)).font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=sum(h.credit_total for h in heads)).font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=sum(h.net for h in heads)).font = Font(bold=True)
        for col in (3, 4, 5):
            ws.cell(row=total_row, column=col).number_format = MONEY

    _autosize(ws, headers)
    ws.freeze_panes = "A2"


def _write_validation(ws, parsed: ParsedStatement) -> None:
    headers = ["Severity", "Type", "Message", "Page", "Line", "Expected", "Actual", "Suggested Fix"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    issues = parsed.validation.all_issues
    for i, issue in enumerate(issues, start=2):
        ws.append([
            issue.severity, issue.issue_type, issue.message,
            issue.page_number, issue.line_number, issue.expected, issue.actual,
            issue.suggested_fix,
        ])
        for col in range(1, len(headers) + 1):
            ws.cell(row=i, column=col).border = BORDER
    if parsed.validation.ocr_confidence is not None:
        ws.cell(row=len(issues) + 3, column=1, value="OCR Confidence").font = Font(bold=True)
        ws.cell(row=len(issues) + 3, column=2, value=parsed.validation.ocr_confidence)
    _autosize(ws, headers)


def _write_insights(ws, parsed: ParsedStatement) -> None:
    ins: InsightsReport = parsed.insights
    row = 1

    def section(title: str) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12, color="1F4E79")
        row += 1

    def kv(label: str, value) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        if isinstance(value, float):
            ws.cell(row=row, column=2).number_format = MONEY
        row += 1

    def insights_block(title: str, items) -> None:
        nonlocal row
        if not items:
            return
        section(title)
        for it in items:
            kv(it.title, it.message)
            if it.detail:
                ws.cell(row=row, column=2, value=it.detail).font = Font(italic=True, size=9)
                row += 1
            row += 1

    insights_block("Income", ins.income)
    insights_block("Spending", ins.spending)
    insights_block("Recurring", ins.recurring)

    if ins.anomalies:
        section("Flags")
        for a in ins.anomalies:
            ws.cell(row=row, column=1, value=a.kind.replace("_", " ")).font = Font(bold=True, color="B45309")
            ws.cell(row=row, column=2, value=a.message)
            row += 1
            if a.suggested_action:
                ws.cell(row=row, column=2, value=f"Action: {a.suggested_action}").font = Font(italic=True, size=9)
                row += 1
            row += 1

    if ins.forecast:
        section("Cash-flow forecast")
        kv("Average monthly income", ins.forecast.avg_monthly_income)
        kv("Average monthly expense", ins.forecast.avg_monthly_expense)
        kv("Summary", ins.forecast.summary)
        header = row
        ws.cell(row=row, column=1, value="Month").font = Font(bold=True)
        ws.cell(row=row, column=2, value="Projected balance").font = Font(bold=True)
        ws.cell(row=row, column=3, value="At risk").font = Font(bold=True)
        row += 1
        for m in ins.forecast.months:
            ws.cell(row=row, column=1, value=m.month)
            ws.cell(row=row, column=2, value=m.projected_balance).number_format = MONEY
            ws.cell(row=row, column=3, value="Yes" if m.at_risk else "")
            row += 1
        if header:
            row += 1

    if ins.tax:
        section("Tax estimate")
        kv("Business expenses", ins.tax.business_expenses)
        kv("Conservative deductible estimate", ins.tax.deductible_estimate)
        kv("Estimated VAT embedded", ins.tax.vat_estimate)
        if ins.tax.business_category_breakdown:
            ws.cell(row=row, column=1, value="By category").font = Font(bold=True)
            for cat, amt in ins.tax.business_category_breakdown.items():
                ws.cell(row=row, column=2, value=f"{cat}: {amt:,.2f}")
                row += 1
            row += 1
        for note in ins.tax.notes:
            ws.cell(row=row, column=1, value=note).font = Font(italic=True, size=9)
            row += 1

    _autosize(ws, ["Metric", "Value"])


def _write_charts(ws, parsed: ParsedStatement) -> None:
    s = parsed.summary
    meta = parsed.meta
    ws.cell(row=1, column=1, value=f"{meta.bank_name} — {meta.account_name}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"{meta.period_start} to {meta.period_end}")

    # Data for charts
    header_row = 4
    ws.cell(row=header_row, column=1, value="Month")
    ws.cell(row=header_row, column=2, value="Credits")
    ws.cell(row=header_row, column=3, value="Debits")
    for i, m in enumerate(s.monthly_cash_flow, start=header_row + 1):
        ws.cell(row=i, column=1, value=m["month"])
        ws.cell(row=i, column=2, value=m["credits"])
        ws.cell(row=i, column=3, value=m["debits"])

    n_months = max(len(s.monthly_cash_flow), 1)
    data_rows = f"{header_row + 1}:{header_row + n_months}"

    # Monthly cash flow (bar chart)
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Monthly Cash Flow"
    bar.y_axis.title = "Amount"
    bar.x_axis.title = "Month"
    bar.add_data(Reference(ws, min_col=2, max_col=3, min_row=header_row, max_row=header_row + n_months), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=header_row + 1, max_row=header_row + n_months))
    ws.add_chart(bar, "E4")

    # Running balance trend (line chart)
    trend_row = header_row + n_months + 2
    ws.cell(row=trend_row, column=1, value="Index")
    ws.cell(row=trend_row, column=2, value="Running Balance")
    balances = [t.balance for t in parsed.transactions if not t.is_beginning_balance and t.balance is not None]
    for i, b in enumerate(balances, start=1):
        ws.cell(row=trend_row + i, column=1, value=i)
        ws.cell(row=trend_row + i, column=2, value=b)
    line = LineChart()
    line.title = "Running Balance Trend"
    line.style = 12
    line.add_data(Reference(ws, min_col=2, min_row=trend_row, max_row=trend_row + len(balances)), titles_from_data=True)
    line.set_categories(Reference(ws, min_col=1, min_row=trend_row + 1, max_row=trend_row + len(balances)))
    ws.add_chart(line, "E20")

    # Income vs Expense (pie chart)
    pie = PieChart()
    pie.title = "Income vs Expense"
    pie.add_data(Reference(ws, min_col=2, max_col=3, min_row=header_row, max_row=header_row + n_months), titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=1, min_row=header_row, max_row=header_row))
    ws.add_chart(pie, "E36")

    # Daily cash flow (bar chart)
    daily_row = trend_row + max(len(balances), 1) + 2
    ws.cell(row=daily_row, column=1, value="Date")
    ws.cell(row=daily_row, column=2, value="Credits")
    ws.cell(row=daily_row, column=3, value="Debits")
    n_daily = len(s.daily_cash_flow)
    for i, d in enumerate(s.daily_cash_flow, start=daily_row + 1):
        ws.cell(row=i, column=1, value=d["date"])
        ws.cell(row=i, column=2, value=d["credits"])
        ws.cell(row=i, column=3, value=d["debits"])
    if n_daily:
        daily = BarChart()
        daily.type = "col"
        daily.style = 10
        daily.title = "Daily Cash Flow"
        daily.y_axis.title = "Amount"
        daily.x_axis.title = "Date"
        daily.add_data(
            Reference(ws, min_col=2, max_col=3, min_row=daily_row, max_row=daily_row + n_daily),
            titles_from_data=True,
        )
        daily.set_categories(Reference(ws, min_col=1, min_row=daily_row + 1, max_row=daily_row + n_daily))
        ws.add_chart(daily, "E52")


def _autosize(ws, headers: list[str]) -> None:
    for idx, header in enumerate(headers, start=1):
        width = min(max(len(str(header)) + 2, 12), 60)
        ws.column_dimensions[get_column_letter(idx)].width = width
