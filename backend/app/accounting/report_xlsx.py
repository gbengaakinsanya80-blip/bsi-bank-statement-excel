"""Excel (XLSX) export of the financial statements (openpyxl).

One sheet per report with the same layout as the PDF export: company name,
report title, period label, then Account / Amount rows with section headers
and subtotals. Amounts carry a real currency number format so the workbook
behaves like something a professional accountant would send.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.accounting.report_pdf import REPORT_TITLES

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="D9E2F3")
TOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
WHITE_BOLD = Font(bold=True, color="FFFFFF", size=11)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = "#,##0.00"


def build_report_xlsx(*, company: dict[str, Any], report_kind: str, data: dict[str, Any]) -> bytes:
    """Render one financial statement report as an Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = REPORT_TITLES[report_kind][:31]
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 18

    r = 1
    ws.cell(row=r, column=1, value=company.get("name") or "Company").font = Font(bold=True, size=14)
    r += 1
    ws.cell(row=r, column=1, value=REPORT_TITLES[report_kind]).font = Font(size=12)
    r += 1
    period = data.get("period_id")
    ws.cell(row=r, column=1, value=period or "Period: All time (entire ledger)")
    ws.cell(row=r, column=1).font = Font(italic=True, color="666666")
    r += 2

    # Column header row.
    for col, label in enumerate(("Account", "Amount"), start=1):
        cell = ws.cell(row=r, column=col, value=label)
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.border = BORDER
    header_row = r
    r += 1

    _populate(ws, r, report_kind, data)

    ws.freeze_panes = f"A{header_row + 1}"
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def _populate(ws, start_row: int, report_kind: str, data: dict[str, Any]) -> None:
    rows: list[tuple[str, str]] = []
    if report_kind == "income-statement":
        _income(rows, data)
    elif report_kind == "balance-sheet":
        _balance(rows, data)
    else:
        _cash_flow(rows, data)

    for i, (label, amount, kind) in enumerate(rows):
        row = start_row + i
        ws.cell(row=row, column=1, value=label)
        amt = ws.cell(row=row, column=2, value=_as_number(amount))
        if amount:
            amt.number_format = MONEY
        amt.alignment = Alignment(horizontal="right")
        if kind == "section":
            ws.cell(row=row, column=1).fill = SECTION_FILL
            ws.cell(row=row, column=1).font = BOLD
            ws.cell(row=row, column=2).fill = SECTION_FILL
        elif kind == "total":
            ws.cell(row=row, column=1).fill = TOTAL_FILL
            ws.cell(row=row, column=1).font = BOLD
            ws.cell(row=row, column=2).fill = TOTAL_FILL
            ws.cell(row=row, column=2).font = BOLD
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = BORDER


def _income(rows: list, data: dict[str, Any]) -> None:
    rows.append(("Revenue", "", "section"))
    for item in data.get("revenue", []):
        rows.append((_acct(item), item["balance"], "line"))
    rows.append(("Total Revenue", data.get("total_revenue", 0), "total"))
    rows.append(("Expenses", "", "section"))
    for item in data.get("expenses", []):
        rows.append((_acct(item), item["balance"], "line"))
    rows.append(("Total Expenses", data.get("total_expenses", 0), "total"))
    rows.append(("Net Profit", data.get("net_profit", 0), "total"))


def _balance(rows: list, data: dict[str, Any]) -> None:
    rows.append(("Assets", "", "section"))
    for item in data.get("assets", []):
        rows.append((_acct(item), item["balance"], "line"))
    rows.append(("Total Assets", data.get("total_assets", 0), "total"))
    rows.append(("Liabilities", "", "section"))
    for item in data.get("liabilities", []):
        rows.append((_acct(item), item["balance"], "line"))
    rows.append(("Total Liabilities", data.get("total_liabilities", 0), "total"))
    rows.append(("Equity", "", "section"))
    for item in data.get("equity", []):
        rows.append((_acct(item), item["balance"], "line"))
    rows.append(("Current Year Profit", data.get("current_year_profit", 0), "line"))
    if float(data.get("balancing_figure", 0)) != 0:
        rows.append(("Balancing Figure", data.get("balancing_figure", 0), "line"))
    rows.append(("Total Equity", data.get("total_equity", 0), "total"))


def _cash_flow(rows: list, data: dict[str, Any]) -> None:
    op = data.get("operating", {})
    rows.append(("Operating Activities", "", "section"))
    rows.append(("Net Profit", op.get("net_profit", 0), "line"))
    for item in op.get("adjustments", []):
        rows.append((f"Adjustment — {_acct(item)}", item.get("change", 0), "line"))
    rows.append(("Net Cash from Operating Activities", op.get("net_cash", 0), "total"))
    inv = data.get("investing", {})
    rows.append(("Investing Activities", "", "section"))
    for item in inv.get("items", []):
        rows.append((_acct(item), item.get("change", 0), "line"))
    rows.append(("Net Cash from Investing Activities", inv.get("net_cash", 0), "total"))
    fin = data.get("financing", {})
    rows.append(("Financing Activities", "", "section"))
    for item in fin.get("items", []):
        rows.append((_acct(item), item.get("change", 0), "line"))
    rows.append(("Net Cash from Financing Activities", fin.get("net_cash", 0), "total"))
    rows.append(("Net Increase in Cash", data.get("net_increase_in_cash", 0), "line"))
    rows.append(("Closing Cash", data.get("closing_cash", 0), "total"))


def _acct(item: dict[str, Any]) -> str:
    code = item.get("code", "")
    name = item.get("name", "")
    return f"{code} — {name}" if code else name


def _as_number(value) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None
